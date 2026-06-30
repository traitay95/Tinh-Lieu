import math
import os
import re
from openpyxl import load_workbook
import streamlit as st

# --- CẤU HÌNH TRANG WEB ---
st.set_page_config(
    page_title="Công Cụ Tính Liều Thuốc - Minh Nhân Professional",
    page_icon="🩺",
    layout="wide",
)

st.title("🩺 Công Cụ Tính Liều Thuốc & Bút Tiêm Insulin")
st.markdown("---")


# --- HÀM ĐỌC DỮ LIỆU EXCEL ---
@st.cache_data(show_spinner=False)
def load_excel_data(uploaded_file):
    try:
        wb = load_workbook(uploaded_file)
        ws = wb.active

        thuoc_dict = {}
        thuoc_quycach_dict = {}

        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:  # Cột A có dữ liệu (Tên thuốc)
                key = str(row[0]).strip()
                # Cột B: Liều dùng
                thuoc_dict[key] = (
                    str(row[1]).strip() if len(row) > 1 and row[1] else ""
                )
                # Cột C: Quy cách
                thuoc_quycach_dict[key] = (
                    str(row[2]).strip() if len(row) > 2 and row[2] else ""
                )

        return thuoc_dict, thuoc_quycach_dict
    except Exception as e:
        st.error(f"❌ Lỗi đọc file Excel: {e}")
        return {}, {}


# --- KHU VỰC CẬP NHẬT FILE EXCEL ---
st.sidebar.header("⚙️ Quản lý Dữ liệu")
uploaded_file = st.sidebar.file_uploader(
    "Tải lên file Excel danh sách thuốc (.xlsx)", type=["xlsx"]
)

if uploaded_file is None:
    default_excel = "danh_sach_thuoc_lieu_dung.xlsx"
    if os.path.exists(default_excel):
        with open(default_excel, "rb") as f:
            thuoc_dict, thuoc_quycach_dict = load_excel_data(f)
        st.sidebar.success("✅ Đang sử dụng file Excel mặc định tại máy.")
    else:
        st.sidebar.warning(
            "⚠️ Không tìm thấy dữ liệu. Vui lòng kéo thả file Excel vào đây!"
        )
        thuoc_dict, thuoc_quycach_dict = {}, {}
else:
    thuoc_dict, thuoc_quycach_dict = load_excel_data(uploaded_file)
    st.sidebar.success("🎉 Đã cập nhật dữ liệu từ file mới thành công!")

danh_sach_thuoc = sorted(thuoc_dict.keys())


# --- GIAO DIỆN CHÍNH: CHIA THÀNH 2 TAB ---
tab1, tab2 = st.tabs(
    ["⚖️ Tính Liều Theo Cân Nặng", "🖊️ Tính Số Lượng Bút Insulin Theo Ngày Kê"]
)

# ==============================================================================
# TAB 1: TÍNH LIỀU THEO CÂN NẶNG
# ==============================================================================
with tab1:
    col1, col2 = st.columns([2, 3])

    with col1:
        st.subheader("Nhập thông tin")
        ten_thuoc_chon = st.selectbox(
            "1. Nhập/Chọn tên thuốc:",
            options=[""] + danh_sach_thuoc,
            key="tab1_thuoc",
        )

        cac_lieu_dung = []
        if ten_thuoc_chon:
            value_b = thuoc_dict.get(ten_thuoc_chon, "")
            if value_b:
                cac_lieu_dung = [
                    v.strip() for v in value_b.split(";") if v.strip()
                ]

        lieu_chon = st.selectbox(
            "2. Chọn liều dùng tương ứng:",
            options=cac_lieu_dung
            if cac_lieu_dung
            else ["(Vui lòng chọn thuốc trước)"],
        )

        can_nang = st.number_input(
            "3. Nhập cân nặng bệnh nhân (kg):",
            min_value=0.0,
            max_value=200.0,
            value=0.0,
            step=0.1,
            format="%.1f",
        )

    with col2:
        st.subheader("📋 Kết quả đề nghị")

        if ten_thuoc_chon and can_nang > 0 and lieu_chon:
            numbers = re.findall(r"[\d.]+", lieu_chon)
            if numbers:
                lieu_dung_so = float(numbers[0])
                tong_lieu = lieu_dung_so * can_nang

                if "ngày" in lieu_chon.lower():
                    tong_lieu_de_nghi = tong_lieu / 2
                else:
                    tong_lieu_de_nghi = tong_lieu

                lieu_theo_tuoi = "tuổi" in lieu_chon.lower()

                if not lieu_theo_tuoi:
                    st.metric(label="Tổng liều tính toán:", value=f"{tong_lieu:.0f} mg")
                else:
                    st.info("💡 Lưu ý: Xem liều dùng chi tiết theo tuổi.")

                # Xử lý quy cách (Cột C)
                quy_cach = thuoc_quycach_dict.get(ten_thuoc_chon, "")
                quy_cach_list = []
                if quy_cach:
                    quy_cach_list = [
                        int(x.strip())
                        for x in quy_cach.split(";")
                        if x.strip().isdigit()
                    ]

                if quy_cach_list:
                    gan_nhat = min(
                        quy_cach_list, key=lambda x: abs(x - tong_lieu_de_nghi)
                    )
                else:
                    gan_nhat = "chưa có cấu hình"

                if not lieu_theo_tuoi:
                    if "ngày" in lieu_chon.lower():
                        st.success(
                            f"👉 **Đề nghị:** Dùng Sáng 1 viên - Chiều 1 viên loại **{gan_nhat} mg**"
                        )
                    else:
                        st.success(
                            f"👉 **Đề nghị:** Một lần dùng là **{gan_nhat} mg**"
                        )
                else:
                    st.warning(f"👉 **Đề nghị:** {lieu_chon}")
        else:
            st.info(
                "Vui lòng chọn thuốc, chọn liều và nhập cân nặng lớn hơn 0 để xem kết quả tính toán."
            )


# ==============================================================================
# TAB 2: TÍNH SỐ LƯỢNG BÚT INSULIN (ĐÃ MẶC ĐỊNH 31 NGÀY)
# ==============================================================================
with tab2:
    col_in1, col_in2 = st.columns([2, 3])

    with col_in1:
        st.subheader("Nhập thông tin kê đơn")

        # Đổi giá trị tham số value thành 31 ngày mặc định
        so_ngay_muon_ke = st.number_input(
            "Nhập số ngày muốn kê đơn (ngày):",
            min_value=1,
            max_value=365,
            value=31,  # <--- Thay đổi ở đây
            step=1,
        )

        tong_lieu_mot_cay_but = st.number_input(
            "Tổng số liều của 1 cây bút tiêm (đơn vị):", value=300, step=50
        )

        lieu_sang = st.number_input(
            "Liều dùng buổi SÁNG (đơn vị):", min_value=0, value=0, step=1
        )
        lieu_chieu = st.number_input(
            "Liều dùng buổi CHIỀU (đơn vị):", min_value=0, value=0, step=1
        )

    with col_in2:
        st.subheader("📆 Kế hoạch sử dụng bút tiêm")

        tong_lieu_ngay = lieu_sang + lieu_chieu

        if tong_lieu_ngay > 0:
            tong_lieu_can_thiet = so_ngay_muon_ke * tong_lieu_ngay
            so_cay_but_tinh_duoc = tong_lieu_can_thiet / tong_lieu_mot_cay_but

            if so_cay_but_tinh_duoc.is_integer():
                so_cay_but_dieu_chinh = so_cay_but_tinh_duoc
                so_ngay_dung = so_ngay_muon_ke
            else:
                if so_ngay_muon_ke >= 90:
                    so_cay_but_dieu_chinh = math.floor(so_cay_but_tinh_duoc)
                    if so_cay_but_dieu_chinh < 1:
                        so_cay_but_dieu_chinh = 1
                else:
                    so_cay_but_dieu_chinh = math.ceil(so_cay_but_tinh_duoc)

                so_ngay_dung = math.floor(
                    (so_cay_but_dieu_chinh * tong_lieu_mot_cay_but)
                    / tong_lieu_ngay
                )

            sang_mot_chieu_mot = so_ngay_dung * 2
            sang_mot_trua_mot_chieu_mot = so_ngay_dung * 3
            sang_hai_chieu_hai = so_ngay_dung * 4
            sang_ba_chieu_ba = so_ngay_dung * 6

            ket_qua_text = (
                f"🩺🖊️ Số cây bút cần kê: {math.floor(so_cay_but_dieu_chinh)} bút\n"
                f"📆 Số ngày dùng thực tế: {so_ngay_dung} ngày (Dự kiến kê: {so_ngay_muon_ke} ngày)\n"
                f"💊💊 Sáng 1 Chiều 1: {sang_mot_chieu_mot} đơn vị\n"
                f"💊💊💊💊 Sáng 2 Chiều 2: {sang_hai_chieu_hai} đơn vị\n"
                f"💊💊💊 Sáng 1 Trưa 1 Chiều 1: {sang_mot_trua_mot_chieu_mot} đơn vị\n"
                f"💊💊💊💊💊💊 Sáng 2 Trưa 2 Chiều 2 [Sáng 3 Chiều 3]: {sang_ba_chieu_ba} đơn vị"
            )

            st.text_area(
                label="Kết quả chi tiết (Có thể sao chép để in):",
                value=ket_qua_text,
                height=220,
            )
        else:
            st.info("Vui lòng nhập liều tiêm sáng hoặc chiều để tính toán.")